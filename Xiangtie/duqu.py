# -*- coding: utf-8 -*-
"""
Created on Thu May 27 11:33:22 2021

@author: Yz Wu
"""
import numpy as np 
from openpyxl import load_workbook
from openpyxl import Workbook

#导入装箱单
packingList = load_workbook('D:\Develop\EndLabel_V2.0\沙滩裤装箱单.xlsx')
#定位装箱单中的工作表


sheet = packingList.worksheets[0]

#导入模板
templateBook = load_workbook('D:\Develop\EndLabel_V2.0\箱贴模板 - Boohoo.xlsx')
#定位模板文件sheet
templateSheet = templateBook.worksheets[0]

#定位箱号列
cells = sheet['G7:''G'+str(sheet.max_row)]
#读取箱号列
values = [[cell.value for cell in columns] for columns in cells]
#降维
b = np.squeeze(values)
print('b=',b)


cel2=sheet['G7:''G'+str(sheet.max_row)]

zuidahang=sheet.max_row

print('zuidahang=',zuidahang)

# print(cells)
# print(cel2)





#定位空值
p=np.where(b==None)
# print('p=',p)
# print('p0=',p[0])

#空值个数
Blank=len(p[0])

print('Blank=',Blank)

#准备一个空集，放n组装箱单
P=[]

####################
if Blank == 0:
    # print('b[0:(p[0][0])]=',b[0:(p[0][0])])
    values+=[[None]]
    b = np.squeeze(values)
    p=np.where(b==None)
    print('b=',b)
    print('p=',p)
    p_0=b[0:(p[0][0])]
    P+=[p_0]
elif Blank == 1:
    p_0=b[0:(p[0][0])]
    P+=[p_0]
    p_l=b[p[0][0]+1:]
    P+=[p_l]
    print('p_l=',p_l)
else:
  
    #产生第一组箱号
    p_0=b[0:(p[0][0])]
    
    # print('p_0=',p_0)
    
    #把第一组箱号存到P中
    P+=[p_0]
    print('P=',P)
    
    #产生第2到n-1组箱号
    for i in range(1,Blank) :
        names='p_'+str(i)
        locals()['p_'+str(i)]=b[p[0][i-1]+1:p[0][i]]
        #把产生的箱号组存到P中
        P+=[b[p[0][i-1]+1:p[0][i]]]
    
    #产生第n组箱号
    p_l=b[p[0][i]+1:]
    
    #把第n组箱号存到P中
    P+=[p_l]
    
    # #print(P[3][0])

print('P=',P)


##############################













##以上，P[n]即为第n组箱号（n=0,1,...n）##
sequence = len(P)+1
print('sequence=',sequence)


# #所有合并的单元格
# merged_ranges = sheet.merged_cell_ranges

# #输入箱号的范围
# for n in range(0,len(P)+1):
#     # 创建一个工作簿对象
#     wb = Workbook()
#     #n的取值要小于数组长度，否则报错
#     if n < len(P):     
#         #循环创建并命名工作表
#         if len(P[n]) !=1:      
#             for i in range(P[n][0],P[n][len(P[n])-1]+1) :
#                 wb.create_sheet(title='箱号'+str(i))
    
#                 Number="箱号"+str(i)
#                 ws=wb[Number]
                         
#                 ##从模板复制
#                 from copy import copy
            
#                 for row in templateSheet.rows:
#                     for cell in row:
#                         new_cell = ws.cell(row=cell.row, column=cell.col_idx,
#                                 value= cell.value)
#                         if cell.has_style:
#                             new_cell.font = copy(cell.font)
#                             new_cell.border = copy(cell.border)
#                             new_cell.fill = copy(cell.fill)
#                             new_cell.number_format = copy(cell.number_format)
#                             new_cell.protection = copy(cell.protection)
#                             new_cell.alignment = copy(cell.alignment)
#                 ##
#                 ## 通过箱号定位，填写当前工作表 
#                 for cellss in cells:
#                     for cell in cellss:
#                         if cell.value ==i:
#                             #箱号i所对应的PO
#                             coor_B='B'+str(cell.row)
#                             #箱号i所对应的款号
#                             coor_C='C'+str(cell.row)
#                             #循环判断PO是否为合并单元格，并把PO号填入当前工作表
#                             for merged_range_B in merged_ranges:
#                                 if sheet[coor_B].coordinate in merged_range_B:
#                                     merged_value = merged_range_B.start_cell.value
#                                     ws['C7']=merged_value
#                                     break
#                                 else:
#                                     PO=sheet['B'+str(cell.row)]
#                                     ws['C7']=PO.value
#                             #循环判断款号是否为合并单元格，并把PO号填入当前工作表
#                             for merged_range_C in merged_ranges:
#                                 if sheet[coor_C].coordinate in merged_range_C:
#                                     merged_value = merged_range_C.start_cell.value
#                                     ws['D10']=merged_value
#                                     break
#                                 else:
#                                     Style=sheet['C'+str(cell.row)]
#                                     ws['D10']=Style.value
                        
#                             ##填写其他信息
#                             Size=sheet['E'+str(cell.row)]
#                             Quantity=sheet['F'+str(cell.row)]
#                             Colour_zh=sheet['K'+str(cell.row)]
#                             Colour_en=sheet['L'+str(cell.row)]
                           
                      
#                             ws['D13']=Size.value
#                             ws['D16']=Colour_en.value
#                             ws['D17']=Colour_zh.value
#                             ws['D19']=Quantity.value
#                             ws['D22']=str(i) + " of " + str(b[-1])
#                             ##
#                             #调整C列的宽度，达到和模板一样
#                             ws.column_dimensions['C'].width = 20
                      
#             #删除多余的"Sheet"页面
#             del wb["Sheet"]      
#             #工作簿名称
#             WorkbookName='箱号'+str(P[n][0])+'-'+str(P[n][len(P[n])-1])+'.xlsx'
#             #保存工作薄
#             wb.save(WorkbookName)
#             # 最后关闭文件
#             wb.close()
           
#         else:
#             #循环创建并命名工作表
#             wb.create_sheet(title='箱号'+str(P[n][0]))
#             Number="箱号"+str(P[n][0])
#             ws=wb[Number]
                         
#             ##从模板复制
#             from copy import copy
        
#             for row in templateSheet.rows:
#                 for cell in row:
#                     new_cell = ws.cell(row=cell.row, column=cell.col_idx,
#                             value= cell.value)
#                     if cell.has_style:
#                         new_cell.font = copy(cell.font)
#                         new_cell.border = copy(cell.border)
#                         new_cell.fill = copy(cell.fill)
#                         new_cell.number_format = copy(cell.number_format)
#                         new_cell.protection = copy(cell.protection)
#                         new_cell.alignment = copy(cell.alignment)
#             ##
#             ## 通过箱号定位，填写当前工作表        
#             for cellss in cells:
#                 for cell in cellss:
#                     if cell.value ==P[n][0]:
                        
           
#                         PO=sheet['B'+str(cell.row)]
#                         Style=sheet['C'+str(cell.row)]
#                         Size=sheet['E'+str(cell.row)]
#                         Quantity=sheet['F'+str(cell.row)]
#                         Colour_zh=sheet['K'+str(cell.row)]
#                         Colour_en=sheet['L'+str(cell.row)]
                
#                         ws['C7']=PO.value
#                         ws['D10']=Style.value
#                         ws['D13']=Size.value
#                         ws['D16']=Colour_en.value
#                         ws['D17']=Colour_zh.value
#                         ws['D19']=Quantity.value
#                         ws['D22']=str(P[n][0]) + " of " + str(b[-1])
#                         ##
#                         #调整C列的宽度，达到和模板一样
#                         ws.column_dimensions['C'].width = 20
#             ##
#             #删除多余的"Sheet"页面
#             del wb["Sheet"]                     
#             #工作簿名称
#             WorkbookName='箱号'+str(P[n][0])+'.xlsx'
#             #保存工作薄
#             wb.save(WorkbookName)
#             # 最后关闭文件
#             wb.close()
# print('')
# print('**************')
# print('箱贴生成完毕')
