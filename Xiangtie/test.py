# -*- coding: utf-8 -*-
"""
Created on Thu May 27 00:07:03 2021

@author: Yz Wu
"""

from openpyxl import load_workbook

#打开模板
templateBook = load_workbook('箱贴模板.xlsx')
#打开目标文件
wb=load_workbook('箱号1-6.xlsx')

#循环复制

for i in range(1,7) :
#定位箱号sheet
    Number="箱号"+str(i)
    ws=wb[Number]
#定位模板文件sheet
    templateSheet = templateBook['1']
    
#复制   
    from copy import copy
    
    for row in templateSheet.rows:
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.col_idx,
                    value= cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
#保存
wb.save('箱号1-6.xlsx')